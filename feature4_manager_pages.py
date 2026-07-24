# -*- coding: utf-8 -*-
"""
Feature 4: 매니저코칭용페이지 생성
명세: specs/feature-4-spec.md

입력  : feature2_manager_view.py의 계산 결과(manager_views) - 별도 외부 입력 없음
동작  : 매니저별로 관련 그룹의 전국 전체 순위표(내 소속 표시 포함)를 담은 엑셀 파일을
        한 명당 하나씩 만든다
출력  : outputs/manager-pages/{매니저이름}.xlsx (매니저당 1개)
        (URL 배포·승인·카카오톡 발송은 이 기능 범위 밖 - feature-4-spec.md '지금은 뺄 것')
"""
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import feature1_engine as f1
import feature2_manager_view as f2

FOLDER = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(FOLDER, "outputs", "manager-pages")

UNSAFE_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')

WON_FORMAT = '#,##0"원"'
PERCENT_FORMAT = "0.0%"
TABLE_HEADERS = ["전국순위", "행사자이름", "소속매니저", "거래처", "채널·품목군 조합", "누적매출", "달성률", "내 소속"]
COLUMN_WIDTHS = [10, 14, 14, 18, 24, 14, 10, 10]
MY_TEAM_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")


def safe_filename_part(text):
    return UNSAFE_FILENAME_CHARS.sub("_", str(text)).strip()


def write_group_block(ws, row, group):
    label = f1.combo_label(group["순위목록"][0]["품목군슬롯"])
    title = f"[{group['채널']} · {label}] 그룹 전국 순위 (총 {len(group['순위목록'])}건)"
    ws.cell(row=row, column=1, value=title).font = Font(bold=True)
    row += 1

    for col, header in enumerate(TABLE_HEADERS, start=1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
    row += 1

    for r in group["순위목록"]:
        ws.cell(row=row, column=1, value=r["순위"])
        ws.cell(row=row, column=2, value=r["행사자이름"])
        ws.cell(row=row, column=3, value=r["매니저"])
        ws.cell(row=row, column=4, value=r["거래처명"])
        ws.cell(row=row, column=5, value=f1.combo_label(r["품목군슬롯"]))
        ws.cell(row=row, column=6, value=r["누적매출"]).number_format = WON_FORMAT
        rate_cell = ws.cell(row=row, column=7, value=r.get("달성률"))
        if r.get("달성률") is not None:
            rate_cell.number_format = PERCENT_FORMAT
        ws.cell(row=row, column=8, value="✅" if r["내소속"] else "")
        if r["내소속"]:
            for col in range(1, len(TABLE_HEADERS) + 1):
                ws.cell(row=row, column=col).fill = MY_TEAM_FILL
        row += 1

    return row + 1  # 그룹 사이 빈 줄


def build_manager_workbook(manager_name, groups):
    wb = Workbook()
    ws = wb.active
    ws.title = "코칭"

    ws["A1"] = f"{manager_name} 매니저 코칭 화면"
    ws["A1"].font = Font(size=14, bold=True)

    row = 3
    for group in groups:
        row = write_group_block(ws, row, group)

    for col_letter, width in zip("ABCDEFGH", COLUMN_WIDTHS):
        ws.column_dimensions[col_letter].width = width

    return wb


def generate_pages(manager_views):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    paths = []
    for manager_name, groups in manager_views.items():
        filename = f"{safe_filename_part(manager_name)}.xlsx"
        path = os.path.join(OUTPUT_DIR, filename)
        build_manager_workbook(manager_name, groups).save(path)
        paths.append(path)
    return paths


def print_intro():
    print("=" * 70)
    print("[매니저코칭용페이지 생성]")
    print("feature2_manager_view.py의 계산 결과를 바탕으로, 매니저 한 명당 엑셀 파일을")
    print("outputs/manager-pages/ 폴더에 하나씩 만듭니다.")


def run(paths=None):
    print_intro()

    result = f1.compute_all(paths)
    manager_views = f2.build_manager_views(result["report"], result["group_rankings"])

    print("=" * 70)
    if not manager_views:
        print("생성할 파일이 없습니다 (유효한 매출 데이터가 없거나 모두 정제 과정에서 제외됨).")
        return []

    generated = generate_pages(manager_views)

    print("[처리 요약]")
    print(f"매니저 {len(generated)}명 파일 생성 완료 -> {OUTPUT_DIR}")
    for path in generated:
        print(f"  - {os.path.basename(path)}")
    return generated


if __name__ == "__main__":
    try:
        run()
    except (FileNotFoundError, ValueError) as e:
        print("=" * 70)
        print("[오류] 실행을 멈췄습니다.")
        print(str(e))
