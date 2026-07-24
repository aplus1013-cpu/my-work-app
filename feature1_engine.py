# -*- coding: utf-8 -*-
"""
Feature 1: 행사자별 매출 요약 계산 엔진
명세: feature-1-spec.md

입력  : 이 폴더에 있는 "[..팀] N월 채널 N주차 유인행사매출 ... 공유(件).xlsx" 파일들
동작  : feature-1-spec.md의 '동작' 순서(파일 검증 -> 컬럼 검증 -> 파싱 -> 결측치 제거
        -> 형식 오류 제거 -> 이상값 제거 -> 행사자 매핑 -> 누적매출/달성률 계산 -> 품목군 그룹핑/순위 산정)
출력  : 행사자별 누적매출/달성률과 거래처별 채널·품목군 조합 전국순위를 콘솔에 출력
        (웹페이지·저장·승인·카카오톡 발송은 이 기능 범위 밖 - feature-1-spec.md '지금은 뺄 것')
"""
import calendar
import glob
import os
import re
from datetime import datetime

# ---------------------------------------------------------------------------
# 설정
# ---------------------------------------------------------------------------

FOLDER = os.path.dirname(os.path.abspath(__file__))
FILE_GLOB = os.path.join(FOLDER, "inputs", "sales", "*유인행사매출*.xlsx")

REQUIRED_COLUMNS = [
    "매니저", "영업담당", "등급", "거래처번호", "거래처명", "2차점ID", "2차점명",
    "행사자사번", "행사자이름", "고용형태", "직무", "마감일", "목표", "총판매금액",
    "달성률", "DayRate", "LineCount", "품목군1", "품목군2", "품목군3",
    "판매수량1", "판매수량2", "판매수량3", "판매금액1", "판매금액2", "판매금액3",
]

# 결측치 검사 대상: 조건부로 비어있을 수 있는 2/3번째 품목군 관련 컬럼과
# 사용하지 않는 DayRate/LineCount는 제외한다.
CORE_REQUIRED_FIELDS = [
    "매니저", "영업담당", "등급", "거래처번호", "거래처명", "2차점ID", "2차점명",
    "행사자사번", "행사자이름", "고용형태", "직무", "마감일", "목표", "총판매금액",
    "달성률", "품목군1", "판매수량1", "판매금액1",
]

SPIKE_MULTIPLIER = 5  # 비정상 급증 기준 배수 (추후 조정 가능, 기획서 참고)

CHANNEL_PATTERN = re.compile(r"\d+월\s+(\S+)\s+\d+주차")
FILENAME_RULE_HINT = (
    '파일명에 "유인행사매출"이 포함되어야 하고, "N월 채널명 N주차" 형태도 들어가야 합니다 '
    '(예: "7월 KAT 3주차 유인행사매출 (~7/15) 공유(件).xlsx").'
)


# ---------------------------------------------------------------------------
# 1~4단계: 파일 검증 / 컬럼 검증 / 파싱
# ---------------------------------------------------------------------------

def extract_channel(filename):
    m = CHANNEL_PATTERN.search(filename)
    if not m:
        raise ValueError(f"{filename}: 파일명에서 채널을 찾을 수 없습니다. {FILENAME_RULE_HINT}")
    return m.group(1)


def is_blank(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def load_and_validate_files(paths):
    """2~4단계: 파일 형식 검증, 필수 컬럼 검증, 표 형태로 파싱."""
    import openpyxl

    all_rows = []
    for path in paths:
        filename = os.path.basename(path)

        # 2. 파일 형식 검증
        if not filename.lower().endswith(".xlsx"):
            raise ValueError(f"지원하지 않는 파일 형식입니다: {filename}")

        channel = extract_channel(filename)

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            raise ValueError(f"{filename}: 헤더 행이 없는 빈 시트입니다")
        header = [str(c.value).strip() if c.value is not None else "" for c in header_row]

        # 3. 26개 필수 컬럼 존재 여부 검증
        missing = [col for col in REQUIRED_COLUMNS if col not in header]
        if missing:
            raise ValueError(f"{filename}: 필수 컬럼 누락 - {missing}")

        col_index = {col: header.index(col) for col in REQUIRED_COLUMNS}

        # 4. 엑셀 데이터를 표 형태로 파싱
        for row_no, cells in enumerate(ws.iter_rows(min_row=2), start=2):
            values = [c.value for c in cells]
            if all(is_blank(v) for v in values):
                continue  # 완전 빈 행은 무시
            record = {col: values[idx] for col, idx in col_index.items()}
            record["_채널"] = channel
            record["_파일"] = filename
            record["_행번호"] = row_no
            all_rows.append(record)

    return all_rows


# ---------------------------------------------------------------------------
# 5단계: 결측치 제거
# ---------------------------------------------------------------------------

def remove_missing(rows):
    """반환되는 dropped는 (행, 사유목록) 형태로, 이상값 제거 결과와 형식을 맞춘다."""
    kept, dropped = [], []
    for r in rows:
        blank_fields = [f for f in CORE_REQUIRED_FIELDS if is_blank(r.get(f))]
        if blank_fields:
            dropped.append((r, [f"결측치({', '.join(blank_fields)} 비어있음)"]))
        else:
            kept.append(r)
    return kept, dropped


# ---------------------------------------------------------------------------
# 5.5단계: 형식 오류 제거 (숫자/날짜 칸에 계산할 수 없는 값이 섞인 행)
# ---------------------------------------------------------------------------

NUMERIC_FIELDS = ["목표", "총판매금액", "판매금액1", "판매금액2", "판매금액3"]


def is_valid_number(value):
    if is_blank(value):
        return True  # 비어있는 값은 결측치 단계에서 이미 처리 대상(필수 필드) 또는 정상적으로 비어있을 수 있는 값
    return isinstance(value, (int, float))


def is_valid_date_value(value):
    try:
        parse_date(value)
        return True
    except (ValueError, TypeError):
        return False


def remove_invalid_format(rows):
    """5.5단계: 금액·목표 칸에 숫자가 아닌 값("미정" 등)이 있거나 마감일이 날짜 형식이
    아니면, 이후 계산이 깨지기 전에 결측치 제거와 같은 방식으로 사유와 함께 걸러낸다."""
    kept, dropped = [], []
    for r in rows:
        problems = [f"형식 오류({f}={r[f]!r}, 숫자 아님)" for f in NUMERIC_FIELDS if not is_valid_number(r[f])]
        if not is_valid_date_value(r["마감일"]):
            problems.append(f"형식 오류(마감일={r['마감일']!r}, 날짜 형식 아님)")
        if problems:
            dropped.append((r, problems))
        else:
            kept.append(r)
    return kept, dropped


# ---------------------------------------------------------------------------
# 6단계: 이상값 제거
# ---------------------------------------------------------------------------

def item_group_slots(r):
    """행 하나의 품목군 조합을 (슬롯번호, 값) 쌍으로 반환 (값이 있는 것만, 1~3개)."""
    slots = []
    for i, field in enumerate(("품목군1", "품목군2", "품목군3"), start=1):
        v = r[field]
        if not is_blank(v):
            slots.append((i, v))
    return slots


def item_group_tuple(r):
    """그룹핑 키로 쓰는 값 전용 튜플. 순서는 무관하고, 들어있는 품목군 값의
    집합(개수·값)이 같으면 같은 그룹으로 취급하도록 정렬해서 비교한다."""
    return tuple(sorted(v for _, v in item_group_slots(r)))


def combo_label(slots):
    """기획서 표기와 맞춘 라벨: 예) '가전(품목군1)', '가전+생활가전(품목군1,2)'."""
    if not slots:
        return ""
    values = "+".join(v for _, v in slots)
    numbers = ",".join(str(i) for i, _ in slots)
    return f"{values}(품목군{numbers})"


def amount_fields_sum(r):
    total = 0
    for f in ("판매금액1", "판매금액2", "판매금액3"):
        v = r[f]
        if not is_blank(v):
            total += v
    return total


def remove_outliers(rows):
    reasons = {}  # id(row) -> 사유 목록

    def flag(r, reason):
        reasons.setdefault(id(r), []).append(reason)

    # (1) 음수 매출
    for r in rows:
        for f in ("판매금액1", "판매금액2", "판매금액3"):
            v = r[f]
            if not is_blank(v) and v < 0:
                flag(r, f"음수 매출({f}={v})")
                break

    # (2) 금액 불일치
    for r in rows:
        if r["총판매금액"] != amount_fields_sum(r):
            flag(r, f"금액 불일치(총판매금액={r['총판매금액']}, 합계={amount_fields_sum(r)})")

    # (3) 비정상 급증: 같은 채널·품목군 조합 그룹 평균 대비 5배 이상
    # (해당 행 자신은 평균 계산에서 제외해야 급증분이 평균을 같이 끌어올리지 않는다)
    groups = {}
    for r in rows:
        key = (r["_채널"], item_group_tuple(r))
        groups.setdefault(key, []).append(r)
    for key, group_rows in groups.items():
        total = sum(r["총판매금액"] for r in group_rows)
        n = len(group_rows)
        if n < 2:
            continue  # 비교할 다른 행이 없으면 급증 여부를 판단할 수 없음
        for r in group_rows:
            others_avg = (total - r["총판매금액"]) / (n - 1)
            if others_avg > 0 and r["총판매금액"] > SPIKE_MULTIPLIER * others_avg:
                flag(r, f"비정상 급증(그룹평균={others_avg:,.0f}, 값={r['총판매금액']:,})")

    # (4) 중복 입력: 거래처번호+행사자사번+마감일+품목군 동일 -> 첫 건만 유지
    seen = set()
    for r in rows:
        dup_key = (r["거래처번호"], r["행사자사번"], str(r["마감일"]), item_group_tuple(r))
        if dup_key in seen:
            flag(r, f"중복 입력({dup_key})")
        else:
            seen.add(dup_key)

    kept = [r for r in rows if id(r) not in reasons]
    dropped = [(r, reasons[id(r)]) for r in rows if id(r) in reasons]
    return kept, dropped


# ---------------------------------------------------------------------------
# 7~11단계: 행사자 단위 매핑, 누적매출/달성률 계산
# ---------------------------------------------------------------------------

def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def compute_period_meta(rows):
    """8~9단계: 최신 마감일 기준 누적일수/월 전체일수 산정.
    (정제 후 유효한 마감일을 가진 행이 하나도 없으면 - 예: 업로드된 파일에 실제 매출 데이터가
    전혀 없는 경우 - 계산할 기준일이 없다는 뜻이므로, 크래시 대신 빈 값으로 반환한다.)"""
    dates = [parse_date(r["마감일"]) for r in rows]
    if not dates:
        return {"최신마감일": None, "누적일수": 0, "해당월전체일수": 0}
    latest_date = max(dates)
    month_start = latest_date.replace(day=1)
    elapsed_days = (latest_date - month_start).days + 1
    days_in_month = calendar.monthrange(latest_date.year, latest_date.month)[1]
    return {"최신마감일": latest_date, "누적일수": elapsed_days, "해당월전체일수": days_in_month}


def compute_employee_summaries(rows, meta):
    """7, 10~11단계: 행사자별 누적매출, 누적일수 환산 달성률 (meta의 누적일수/월 전체일수 사용)."""
    elapsed_days = meta["누적일수"]
    days_in_month = meta["해당월전체일수"]

    by_employee = {}
    for r in rows:
        key = (r["행사자사번"], r["행사자이름"])
        by_employee.setdefault(key, []).append(r)

    summaries = {}
    for (emp_no, emp_name), emp_rows in by_employee.items():
        cumulative_sales = sum(r["총판매금액"] for r in emp_rows)
        monthly_target = emp_rows[0]["목표"]  # 행사자당 월 목표는 동일한 값이 반복 기재됨
        manager_name = emp_rows[0]["매니저"]  # 행사자당 매니저도 동일한 값이 반복 기재됨
        prorated_target = monthly_target * elapsed_days / days_in_month if days_in_month else 0
        achievement_rate = cumulative_sales / prorated_target if prorated_target else 0
        summaries[(emp_no, emp_name)] = {
            "행사자이름": emp_name,
            "행사자사번": emp_no,
            "매니저": manager_name,
            "누적매출": cumulative_sales,
            "월목표": monthly_target,
            "환산목표": prorated_target,
            "달성률": achievement_rate,
        }
    return summaries


# ---------------------------------------------------------------------------
# 12~16단계: 채널·품목군 조합 그룹핑 및 전국순위 산정
# ---------------------------------------------------------------------------

def compute_store_rankings(rows):
    """
    12단계: 매출 건(=행사자·거래처 조합)의 채널/품목군 조합 그룹 키 구성
    13단계: 같은 채널 내 품목군 조합이 일치하는 항목끼리 그룹핑
    14단계: 그룹별 일평균매출(누적매출 ÷ 근무일수) 기준 순위 산정
    16단계: 매출 건(거래처·품목군 조합)별 전국순위 매칭

    순위는 누적 총판매금액 원값이 아니라 "근무일수로 나눈 일평균매출"로 매긴다.
    행사 진행 일수(누적일수)가 사람마다 달라, 원값 기준으로 매기면 같은 거래처를
    더 여러 번 뛴 사람이 단지 그 이유만으로 유리해지기 때문이다.
    """
    # 행사자-거래처 단위로 총판매금액 합산 + 해당 거래처에서 뛴 고유 마감일 수(근무일수) 집계
    store_entries = {}
    for r in rows:
        key = (r["행사자사번"], r["행사자이름"], r["거래처번호"], r["거래처명"])
        entry = store_entries.setdefault(key, {
            "매니저": r["매니저"],
            "채널": r["_채널"],
            "품목군조합": item_group_tuple(r),
            "품목군슬롯": item_group_slots(r),
            "누적매출": 0,
            "근무일": set(),
        })
        entry["누적매출"] += r["총판매금액"]
        entry["근무일"].add(str(r["마감일"]))

    for entry in store_entries.values():
        entry["근무일수"] = len(entry["근무일"])
        entry["일평균매출"] = entry["누적매출"] / entry["근무일수"]

    # 채널 + 품목군 조합 그룹별로 묶기
    groups = {}
    for key, entry in store_entries.items():
        group_key = (entry["채널"], entry["품목군조합"])
        groups.setdefault(group_key, []).append((key, entry))

    # 그룹별 순위 산정 (일평균매출 내림차순)
    # 순위는 매출 건(거래처) 단위로 매기되, "그룹 인원수"는 같은 행사자가 같은 그룹에
    # 거래처 여러 곳으로 들어가 있어도 한 명으로만 센다 (엔트리 수로 세면 인원수가 부풀려짐).
    #
    # rankings: 행사자·거래처 키 -> 본인 순위/인원수만 담은 값 (Feature 1 자체 출력용)
    # group_rankings: 그룹 키 -> 그 그룹에 속한 전국 전체 행사자 목록을 순위순으로 담은 값
    #                 (다른 매니저 소속 포함 — Feature 2의 매니저 화면이 이 전체 명단을 그대로 씀)
    rankings = {}
    group_rankings = {}
    for group_key, members in groups.items():
        ranked = sorted(members, key=lambda kv: kv[1]["일평균매출"], reverse=True)
        distinct_headcount = len({key[0] for key, _ in members})  # 행사자사번 기준 dedup

        group_list = []
        for rank, (key, entry) in enumerate(ranked, start=1):
            emp_no, emp_name, _store_no, store_name = key
            rankings[key] = {
                "채널": entry["채널"],
                "품목군조합": entry["품목군조합"],
                "품목군슬롯": entry["품목군슬롯"],
                "누적매출": entry["누적매출"],
                "근무일수": entry["근무일수"],
                "일평균매출": entry["일평균매출"],
                "순위": rank,
                "그룹인원수": distinct_headcount,
            }
            group_list.append({
                "순위": rank,
                "행사자사번": emp_no,
                "행사자이름": emp_name,
                "매니저": entry["매니저"],
                "거래처명": store_name,
                "채널": entry["채널"],
                "품목군조합": entry["품목군조합"],
                "품목군슬롯": entry["품목군슬롯"],
                "누적매출": entry["누적매출"],
                "근무일수": entry["근무일수"],
                "일평균매출": entry["일평균매출"],
            })
        group_rankings[group_key] = group_list

    return rankings, group_rankings


# ---------------------------------------------------------------------------
# 17단계: 행사자별 출력 데이터 구성 + 콘솔 출력
# ---------------------------------------------------------------------------

def build_report(employee_summaries, store_rankings):
    by_employee = {}
    for (emp_no, emp_name, store_no, store_name), ranking in store_rankings.items():
        by_employee.setdefault((emp_no, emp_name), []).append({
            "거래처명": store_name,
            **ranking,
        })

    report = []
    for key, summary in employee_summaries.items():
        stores = by_employee.get(key, [])
        stores.sort(key=lambda s: s["순위"])
        report.append({**summary, "거래처목록": stores})

    report.sort(key=lambda e: e["행사자이름"])
    return report


def format_won(amount):
    return f"{amount:,.0f}원"


def print_report(report, meta):
    print("=" * 70)
    if meta["최신마감일"] is None:
        print("[매출 요약 계산 결과] 유효한 매출 데이터가 없어 계산할 수 없습니다.")
        return
    print(f"[매출 요약 계산 결과] 기준일: {meta['최신마감일']} "
          f"(누적일수 {meta['누적일수']}일 / 해당월 전체일수 {meta['해당월전체일수']}일)")
    print("=" * 70)
    if not report:
        print("\n결과에 포함된 행사자가 없습니다 (모든 매출 행이 정제 과정에서 제외됨).")
        return
    for e in report:
        print(f"\n{e['행사자이름']} 님 (매니저: {e['매니저']}), 이번 달 누적 매출 {format_won(e['누적매출'])} "
              f"(누적일수 기준 목표 대비 {e['달성률'] * 100:.1f}% 달성)")
        for s in e["거래처목록"]:
            label = combo_label(s["품목군슬롯"])
            print(f"  - {s['거래처명']}: {s['채널']}채널 · {label} 그룹 "
                  f"전국 {s['그룹인원수']}명 중 {s['순위']}위 "
                  f"(근무 {s['근무일수']}일 · 일평균 {format_won(s['일평균매출'])} · "
                  f"누적 {format_won(s['누적매출'])})")


def print_cleaning_log(dropped_missing, dropped_format, dropped_outliers, total_raw):
    print("=" * 70)
    print("[데이터 정제 로그]")
    print(f"원본 행 수: {total_raw}건")
    print(f"결측치 제거: {len(dropped_missing)}건")
    for r, reasons in dropped_missing:
        print(f"  - {r['_파일']} {r['_행번호']}행: 행사자={r.get('행사자이름')}, "
              f"거래처={r.get('거래처명')} -> {', '.join(reasons)}")
    print(f"형식 오류 제거: {len(dropped_format)}건")
    for r, reasons in dropped_format:
        print(f"  - {r['_파일']} {r['_행번호']}행: 행사자={r.get('행사자이름')}, "
              f"거래처={r.get('거래처명')} -> {', '.join(reasons)}")
    print(f"이상값 제거: {len(dropped_outliers)}건")
    for r, reasons in dropped_outliers:
        print(f"  - {r['_파일']} {r['_행번호']}행: 행사자={r.get('행사자이름')}, "
              f"거래처={r.get('거래처명')} -> {', '.join(reasons)}")
    print(f"정제 후 남은 행 수: {total_raw - len(dropped_missing) - len(dropped_format) - len(dropped_outliers)}건")


def compute_excluded_employees(all_dropped, cleaned_rows):
    """모든 행이 제거되어 최종 결과에 아예 나타나지 않는 행사자를, 사유와 함께 모은다."""
    survived = {(r["행사자사번"], r["행사자이름"]) for r in cleaned_rows}
    reasons_by_employee = {}
    for r, reasons in all_dropped:
        key = (r["행사자사번"], r["행사자이름"])
        reasons_by_employee.setdefault(key, []).extend(
            f"{r.get('거래처명')} {r.get('마감일')} - {reason}" for reason in reasons
        )
    return {k: v for k, v in reasons_by_employee.items() if k not in survived}


def print_excluded_employees(excluded):
    if not excluded:
        return

    print("=" * 70)
    print("[결과에서 완전히 제외된 행사자]")
    for (emp_no, emp_name), reasons in sorted(excluded.items(), key=lambda kv: kv[0][1]):
        print(f"- {emp_name}({emp_no}): 유효한 매출 행이 하나도 남지 않아 제외됨")
        for reason in reasons:
            print(f"    · {reason}")


def print_summary(paths, result, excluded_count):
    """전체 처리 규모를 한눈에 보여주는 요약 (상세 로그는 아래 섹션들에서 확인)."""
    channels = sorted({extract_channel(os.path.basename(p)) for p in paths})
    cleaned_count = len(result["cleaned_rows"])

    print("=" * 70)
    print("[처리 요약]")
    print(f"입력 파일 {len(paths)}개 (채널: {', '.join(channels)})")
    print(f"원본 {result['raw_row_count']}건 -> 정제 후 {cleaned_count}건 "
          f"(결측치 {len(result['dropped_missing'])}건, 형식 오류 {len(result['dropped_format'])}건, "
          f"이상값 {len(result['dropped_outliers'])}건 제외)")
    print(f"행사자 {len(result['report'])}명 결과 포함, {excluded_count}명 완전 제외")


# ---------------------------------------------------------------------------
# 실행
# ---------------------------------------------------------------------------

def attach_achievement_rate(group_rankings, employee_summaries):
    """그룹 순위 행마다 해당 행사자의 목표대비 달성률을 붙인다 (달성률은 원래 행사자당 1개
    값으로만 존재하는데, 매니저코칭페이지 등 후속 기능은 순위 행 단위로 이 값이 필요하다)."""
    for group_list in group_rankings.values():
        for row in group_list:
            summary = employee_summaries.get((row["행사자사번"], row["행사자이름"]))
            row["달성률"] = summary["달성률"] if summary else None


def compute_all(paths=None):
    """콘솔 출력 없이 전체 계산 파이프라인만 실행. 다른 기능(Feature 2 등)이 이 결과를
    그대로 가져다 쓸 수 있도록, group_rankings(그룹별 전국 전체 명단)까지 함께 반환한다."""
    paths = paths or sorted(glob.glob(FILE_GLOB))
    if not paths:
        raise FileNotFoundError(f"inputs/sales 폴더에 매출 엑셀 파일을 넣어주세요. {FILENAME_RULE_HINT}")

    raw_rows = load_and_validate_files(paths)
    after_missing, dropped_missing = remove_missing(raw_rows)
    after_format, dropped_format = remove_invalid_format(after_missing)
    cleaned_rows, dropped_outliers = remove_outliers(after_format)

    meta = compute_period_meta(after_format)
    employee_summaries = compute_employee_summaries(cleaned_rows, meta)
    store_rankings, group_rankings = compute_store_rankings(cleaned_rows)
    attach_achievement_rate(group_rankings, employee_summaries)
    report = build_report(employee_summaries, store_rankings)

    return {
        "paths": paths,
        "report": report,
        "meta": meta,
        "group_rankings": group_rankings,
        "raw_row_count": len(raw_rows),
        "dropped_missing": dropped_missing,
        "dropped_format": dropped_format,
        "dropped_outliers": dropped_outliers,
        "cleaned_rows": cleaned_rows,
    }


def print_intro():
    print("=" * 70)
    print("[판촉사원 매출 요약 엔진]")
    print("inputs/sales 폴더의 매출 엑셀 파일을 읽어, 행사자별 누적매출·목표대비 달성률과")
    print("거래처별 전국순위를 계산합니다.")
    print(f"파일명 규칙: {FILENAME_RULE_HINT}")


def run(paths=None):
    print_intro()

    result = compute_all(paths)
    all_dropped = result["dropped_missing"] + result["dropped_format"] + result["dropped_outliers"]
    excluded = compute_excluded_employees(all_dropped, result["cleaned_rows"])

    print_summary(result["paths"], result, len(excluded))
    print_cleaning_log(result["dropped_missing"], result["dropped_format"], result["dropped_outliers"], result["raw_row_count"])
    print_excluded_employees(excluded)
    print_report(result["report"], result["meta"])

    return result["report"], result["meta"]


if __name__ == "__main__":
    try:
        run()
    except (FileNotFoundError, ValueError) as e:
        print("=" * 70)
        print("[오류] 실행을 멈췄습니다.")
        print(str(e))
