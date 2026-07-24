# -*- coding: utf-8 -*-
"""
Feature 3: 행사자별 요약페이지 생성
명세: specs/feature-3-spec.md

입력  : feature1_engine.py의 계산 결과(report) - 별도 외부 입력 없음
동작  : 행사자별로 누적매출·목표대비 달성률·거래처별 전국순위를 담은 엑셀 파일을
        한 명당 하나씩 만든다
출력  : outputs/employee-pages/{행사자사번}_{행사자이름}.xlsx (행사자당 1개)
        (URL 배포·승인·카카오톡 발송은 이 기능 범위 밖 - feature-3-spec.md '지금은 뺄 것')
"""
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font

import feature1_engine as f1

FOLDER = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(FOLDER, "outputs", "employee-pages")

UNSAFE_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')

WON_FORMAT = '#,##0"원"'
PERCENT_FORMAT = "0.0%"
TABLE_HEADERS = ["거래처", "채널", "품목군 조합", "전국순위", "그룹인원수", "근무일수", "일평균매출", "누적매출"]
COLUMN_WIDTHS = [18, 8, 22, 10, 10, 10, 14, 14]


def safe_filename_part(text):
    return UNSAFE_FILENAME_CHARS.sub("_", str(text)).strip()


def build_employee_workbook(employee):
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"

    ws["A1"] = f"{employee['행사자이름']} 님"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "이번 달 누적 매출"
    ws["B2"] = employee["누적매출"]
    ws["B2"].number_format = WON_FORMAT
    ws["A3"] = "누적일수 기준 목표 대비 달성률"
    ws["B3"] = employee["달성률"]
    ws["B3"].number_format = PERCENT_FORMAT

    header_row = 5
    for col, header in enumerate(TABLE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)

    for i, s in enumerate(employee["거래처목록"], start=1):
        row = header_row + i
        ws.cell(row=row, column=1, value=s["거래처명"])
        ws.cell(row=row, column=2, value=s["채널"])
        ws.cell(row=row, column=3, value=f1.combo_label(s["품목군슬롯"]))
        ws.cell(row=row, column=4, value=s["순위"])
        ws.cell(row=row, column=5, value=s["그룹인원수"])
        ws.cell(row=row, column=6, value=s["근무일수"])
        ws.cell(row=row, column=7, value=s["일평균매출"]).number_format = WON_FORMAT
        ws.cell(row=row, column=8, value=s["누적매출"]).number_format = WON_FORMAT

    for col_letter, width in zip("ABCDEFGH", COLUMN_WIDTHS):
        ws.column_dimensions[col_letter].width = width

    return wb


def generate_pages(report):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    paths = []
    for employee in report:
        filename = f"{safe_filename_part(employee['행사자사번'])}_{safe_filename_part(employee['행사자이름'])}.xlsx"
        path = os.path.join(OUTPUT_DIR, filename)
        build_employee_workbook(employee).save(path)
        paths.append(path)
    return paths


def print_intro():
    print("=" * 70)
    print("[행사자별 요약페이지 생성]")
    print("feature1_engine.py의 계산 결과를 바탕으로, 행사자 한 명당 엑셀 파일을")
    print("outputs/employee-pages/ 폴더에 하나씩 만듭니다.")


def run(paths=None):
    print_intro()

    result = f1.compute_all(paths)
    report = result["report"]

    print("=" * 70)
    if not report:
        print("생성할 파일이 없습니다 (유효한 매출 데이터가 없거나 모두 정제 과정에서 제외됨).")
        return []

    generated = generate_pages(report)

    print("[처리 요약]")
    print(f"행사자 {len(generated)}명 파일 생성 완료 -> {OUTPUT_DIR}")
    for path in generated:
        print(f"  - {os.path.basename(path)}")
    return generated


if __name__ == "__main__":
    try:
        run()
    except (FileNotFoundError, ValueError) as e:
        print("=" * 70)
        print("[오류] 실행을 멈췄습니다.")
        print(str(e))
